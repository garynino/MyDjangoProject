# Create your views here.
import csv
import json
import time
import urllib.parse

import openpyxl
from bs4 import BeautifulSoup
from django.core.files.base import ContentFile
from django.db import connection
from django.shortcuts import render

from testapp1.models import *
import xml.etree.ElementTree as ET
import zipfile
from django.http import JsonResponse

from zipfile import Path
from django.http import HttpResponse
from django.conf import settings
from django.core.files.storage import default_storage

from openpyxl.utils import get_column_letter

def upload_page(request):
    return render(request, "upload.html")  # Adjust if needed

def export_csv(request):

    #

    print("export_csv view triggered") # used to make sure view function is being called

    necessary_keys_dict = {
        "welcome_course": {"textbook_id": []},
    }

    def make_new_sheet(wb, query):
        #
        cursor = connection.cursor()
        cursor.execute(query) # cursor holds result from cursor.execute() query
        table_name = query.split("FROM")[1].strip()

        rows = cursor.fetchall()
        columns = [col[0] for col in cursor.description]
        sheet = wb.create_sheet(title=table_name)

        # Write headers
        for col_num, column_title in enumerate(columns, 1):
            col_letter = get_column_letter(col_num)
            sheet[f'{col_letter}1'] = column_title

        # Write data rows
        for row_num, row in enumerate(rows, 2):  # Start from row 2
            for col_num, cell_value in enumerate(row, 1):
                col_letter = get_column_letter(col_num)
                sheet[f'{col_letter}{row_num}'] = cell_value

        # end of make_new_sheet inner-function

    # this function finds out whether the given record/entry is missing in the sheet by using record ID
    def is_record_missing(id_to_check, sheet):
        did_not_find = True
        column_of_cells = sheet['A'] # this grabs a column of cells from an Excel worksheet
        for cell in column_of_cells:
            if cell.value == id_to_check:
                did_not_find = False
                return did_not_find

        return did_not_find

    def fill_out_sheet(sheet, query, lists_of_ids_dict, ids_to_grab_list):
        cursor = connection.cursor()
        cursor.execute(query)  # cursor holds result from cursor.execute() query
        # grabs all rows that cursor is holding. each list object is a tuple. doesn't grab column-name row
        rows = cursor.fetchall()
        # this extracts column names from table metadata
        # cursor.description is a list of tuples that contain table metadata
        # column_info designates the tuple object that contains column info
        # the first object in the column metadata-tuple is the column name; column_info[0]
        column_name_list = [column_info[0] for column_info in cursor.description]

        # this block enters the column names from database into the first row of Excel sheet
        # uses the form:        for index, item in enumerate(my_list, number):
        # number is where you start enumerating from. this enumeration starts from 1 (inclusive)
        column_list_pair_dict = {} # BE CAREFUL! every column is actually column-1 for use with row list indexing
        for cell_column, column_name in enumerate(column_name_list, 1):
            cell_column_letter = get_column_letter(cell_column)  # get_column_letter maps number to letter
            sheet[f'{cell_column_letter}1'] = column_name

            if column_name in lists_of_ids_dict:
                row_index = cell_column - 1
                column_list_pair_dict[row_index] = lists_of_ids_dict.get(column_name) # creates int-list pair

        current_records_in_file = 0
        for column_to_check, list_with_ids in column_list_pair_dict.items():
            # this block adds each desired database record to the Excel sheet
            #current_records_in_file = 0 # commented out because i might need to put it back later
            for row in rows:  # for every record/entry in list of records ...
                if row[column_to_check] in list_with_ids and is_record_missing(row[0], sheet):  # check if ID in given ID-list
                    for cell_column, field_value in enumerate(row, 1):
                        cell_column_letter = get_column_letter(cell_column)
                        sheet[f'{cell_column_letter}{(current_records_in_file + 2)}'] = field_value
                    current_records_in_file += 1

        dict_to_return = None
        if ids_to_grab_list is not None:
            dict_to_return = {} # (CHANGED) was list of lists. now, dictionary of lists.
            #current_number_of_lists = 0
            for index, column_id in enumerate(ids_to_grab_list, 0):
                # dict_to_return.append([]) # (BEFORE). useless right now. might need later
                desired_id = ids_to_grab_list[index]
                dict_to_return[desired_id]= [] # (AFTER)
                desired_id_column_letter = None # this is just to initialize and keep value outside the loop
                for first_row_cell in sheet[1]:  # Gets all cells in row 1 of sheet and iterates
                    if first_row_cell.value == desired_id: # if we found the column of the desired id
                        desired_id_column_letter = get_column_letter(first_row_cell.column)
                alleged_cell_column = sheet[desired_id_column_letter] # get entire column of cells
                i = 1
                for cell_in_column in alleged_cell_column:
                    if i > 1:
                        dict_to_return.get(desired_id).append(cell_in_column.value)
                    i += 1

        return dict_to_return

    if request.method == "POST":
        try: # json.loads() will cause an error if the json is invalid or empty
            data = json.loads(request.body) # parses json and creates/saves into a dictionary
        except json.JSONDecodeError: # might change to "json.decoder.JSONDecodeError"
            print('Error. JSON is invalid or empty.')
            return JsonResponse({'error': 'Invalid or empty JSON provided'}, status=400)

        # gets lists from list dictionary. all of these are the ACTUAL IDs from the database (1st column)
        course_id_list = data.get('course', [])
        test_id_list = data.get('test', [])
        question_id_list = data.get('questions', [])
        type_of_export = data.get('typeOfExport', [])

        course_id_list = list(set(course_id_list))
        test_id_list = list(set(test_id_list))
        question_id_list = list(set(question_id_list))

        try: # tries to convert every list into a list of integers (except for type_of_export)
            course_id_list = [int(course_id) for course_id in course_id_list]
            test_id_list = [int(test_id) for test_id in test_id_list]
            question_id_list = [int(question_id) for question_id in question_id_list]
        except ValueError:
            print('ID with NON-numeric ID-value provided')
            return JsonResponse({'error': 'ID with NON-numeric ID-value provided'}, status=400)

        try:
            export_type = type_of_export[0]
        except IndexError:
            print('Export type not given')
            return JsonResponse({'error': 'Export type not provided'}, status=400)

        wb = openpyxl.Workbook() # creates the Workbook container object (Excel file)
        wb.remove(wb.active)  # Remove the default blank sheet

        if export_type == 'entire':
            print(export_type) # placeholder

        elif export_type == 'course':
            print(course_id_list)
            
            if not course_id_list: # if given course list exists but is empty
                print('No courses given to export')
                return JsonResponse({'error': 'No courses given to export'}, status=400)

            sheet = wb.create_sheet(title="welcome_course")
            query = "SELECT * FROM welcome_course"

            """ going to use these in a sec ...
            table_name = 'welcome_course'
            query = f"SELECT * FROM `{table_name}` LIMIT 100"
            query = f"SELECT * FROM `{table_name}`"
            """

            id_dict = {}
            id_dict['id'] = course_id_list # this makes a key-list pair entry in a dictionary
            needed_ids_list = ["textbook_id"]
            result_dict = fill_out_sheet(sheet, query, id_dict, needed_ids_list)
            print(result_dict)


        elif export_type == 'test':
            print(test_id_list)

            if test_id_list:
                print('') # placeholder
            else:
                print('No tests given to export')
                return JsonResponse({'error': 'No tests given to export'}, status=400)

        elif export_type == 'questions':
            print(question_id_list)

            if question_id_list:
                print('') # placeholder
            else:
                print('No questions given to export')
                return JsonResponse({'error': 'No questions given to export'}, status=400)

        else:
            print('Invalid export type given')
            return JsonResponse({'error': 'Invalid export type provided'}, status=400)

        print("Final sheets in workbook:", wb.sheetnames)

        # Save to a BytesIO stream instead of a file
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Prepare the response
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=exported_data.xlsx'

    print("did it get here???") # used for debugging

    return response

#


def parse_qti_xml(request):
    """
    Parses a QTI zip file and saves extracted data to the database.
    This supports QTI version 1.2 only.
    """
    start_time = time.perf_counter()

    class ImageDataPair:
        def __init__(self, raw_image_data, actual_image_name):
            self.raw_image_data = raw_image_data
            self.actual_image_name = actual_image_name

    # Function to remove namespaces
    def remove_namespace(given_tree):
        for elem in given_tree.iter():
            if "}" in elem.tag:
                elem.tag = elem.tag.split("}")[-1]

    # creates a new question record/entry
    def create_question(g_course, g_q_type, g_q_text, g_points):
        temp_question_instance = Question.objects.create(
            course=g_course,
            # this is because, logically, when questions/tests are uploaded to a course, are they not part of it?
            qtype=g_q_type,
            text=g_q_text,
            score=g_points
        )

        # checks if user is logged in
        if request.user.is_authenticated:
            temp_question_instance.author = request.user  # sets to the current user
            temp_question_instance.save()

        return temp_question_instance

    def check_embedded_graphic(text_q):
        if text_q is None:
            return None
        # Parse the HTML using BeautifulSoup4 library
        soup = BeautifulSoup(text_q, 'html.parser')
        # Find the <img> tag
        img_element = soup.find('img')
        found_the_image = False
        if img_element:  # if we find an embedded image
            temp_file_path1 = img_element.get('src')  # gets the src attribute of img element
            url_encoded_path = temp_file_path1[18:]  # length of "$IMS-CC-FILEBASE$/" is 18
            decoded_path = urllib.parse.unquote(url_encoded_path)  # this decodes URL-encoded path

            for potential_image_name in filename_list:  # this assumes an outer function has: filename_list = zip_ref.namelist()
                if potential_image_name.endswith(decoded_path):
                    found_the_image = True
                    my_image_name = img_element.get('alt')
                    with zip_ref.open(potential_image_name) as desired_img_file:  # open the image file
                        img_data = desired_img_file.read()  # this is the raw image data
                        data_to_return = ImageDataPair(img_data, my_image_name)
            if not found_the_image:
                print('Desired image not found')  # used for debugging
        if img_element is None or found_the_image == False:
            return None
        else:
            return data_to_return

    def parse_just_xml(meta_path, non_meta_path, the_course):

        print(f"processing file: {meta_path}")
        # path to metadata file
        xml_file_path = meta_path
        tree = ET.parse(xml_file_path)
        root = tree.getroot()
        remove_namespace(root)

        cover_instructions_text = root.find('.//description').text

        print(f"processing file: {non_meta_path}")
        # Path to the questions file
        xml_file_path = non_meta_path
        tree = ET.parse(xml_file_path)
        root = tree.getroot()

        namespace = {"ns": "http://www.imsglobal.org/xsd/ims_qtiasiv1p2"}

        # Remove namespaces
        remove_namespace(root)
        # Find the 'assessment' element
        my_tag = "assessment"
        node = root.find(my_tag)

        if node is None:
            return JsonResponse({"error": f"Element '{my_tag}' not found in XML!"}, status=400)

        # Extract 'ident' and 'title' attribute from the element that node represents
        the_test_title = node.get("title")  # test name
        test_identifier = node.get("ident")

        # Create a new Test record
        test_instance = Test.objects.create(
            course=the_course,
            textbook=the_course.textbook,
            name=the_test_title
        )
        test_part_instance = TestPart.objects.create(
            test=test_instance
        )
        number_of_sections = 0

        for section in root.findall(".//section"):
            number_of_sections = number_of_sections + 1
            test_section_instance = TestSection.objects.create(
                part=test_part_instance,
                section_number=number_of_sections
            )

            for item in section.findall(".//item"):

                # all useful metadata fields are found in the fieldentry elements under itemmetadata
                node = item.find('itemmetadata')
                qti_metadata_fields = node.findall(".//fieldentry")

                node = item.find('presentation')

                temp_node = node.find('material')  # possibly redundant statement
                temp_node = temp_node.find(".//mattext")
                # question_text_field contains the question prompt text
                question_text_field = temp_node.text

                # each response has an ID represented as "ident" in the XML
                # the ID is used to know what the correct response is
                correct_answer_ident = None

                correct_answer_ident = None

                the_question_type = qti_metadata_fields[0].text
                max_points_for_question = float(qti_metadata_fields[1].text)

                # node should currently be already = element w 'presentation' tag
                # MultiChoice & TF might be able to be combined. For now, they are separate
                if the_question_type == 'multiple_choice_question':
                    #
                    node = node.find('.//response_lid')
                    answer_choices_dict = {}
                    for response_label_elem in node.findall('.//response_label'):
                        response_ident = response_label_elem.get('ident')
                        response_text = response_label_elem.find('.//mattext').text
                        answer_choices_dict[response_ident] = response_text

                    node = item.find('resprocessing')
                    for respcondition_elem in node.findall('.//respcondition'):
                        if respcondition_elem.get('continue') == "No":
                            temp_node = respcondition_elem.find('.//varequal')
                            correct_answer_ident = temp_node.text

                    # this creates a question record in database
                    question_instance = create_question(the_course, the_question_type, question_text_field,
                                                        max_points_for_question)
                    # this block of code checks for an embedded graphic in text, then saves it to field
                    # is there is a graphic. otherwise it does nothing.
                    image_data_pair = check_embedded_graphic(question_text_field)
                    if image_data_pair is not None:
                        # Save the image to the img field, then update the record/entry
                        question_instance.img.save(image_data_pair.actual_image_name,
                                                   ContentFile(image_data_pair.raw_image_data))
                        question_instance.save()
                        print(f"{question_instance.img.url}")

                        # Parse the HTML using BeautifulSoup4 library
                        html_obj = BeautifulSoup(question_text_field, 'html.parser')
                        # Find the element with "img" tag
                        my_img_element = html_obj.find('img')
                        my_img_element['src'] = question_instance.img.url  # change src attribute
                        question_text_field = str(html_obj)  # save html as string
                        question_instance.text = question_text_field  # update text field
                        question_instance.save()  # save/update entry in database

                    testquestion_instance = TestQuestion.objects.create(
                        test=test_instance,
                        question=question_instance,
                        assigned_points=max_points_for_question,
                        section=test_section_instance
                    )

                    for key, value in answer_choices_dict.items():
                        if key == correct_answer_ident:
                            question_instance.answer = value
                            temp_img_data_pair = check_embedded_graphic(value)
                            if temp_img_data_pair is not None:
                                question_instance.ansimg.save(temp_img_data_pair.actual_image_name,
                                                              ContentFile(temp_img_data_pair.raw_image_data))

                                # Parse the HTML using BeautifulSoup4 library
                                html_obj = BeautifulSoup(value, 'html.parser')
                                # Find the element with "img" tag
                                my_img_element = html_obj.find('img')
                                my_img_element['src'] = question_instance.ansimg.url  # change src attribute
                                value = str(html_obj)  # save html as string
                                question_instance.answer = value  # update answer text field
                                question_instance.save()  # save/update entry in database

                            question_instance.save()
                        else:
                            options_instance = Options.objects.create(
                                question=question_instance,
                                text=value
                            )
                            temp_img_data_pair = check_embedded_graphic(value)
                            if temp_img_data_pair is not None:
                                options_instance.image.save(temp_img_data_pair.actual_image_name,
                                                            ContentFile(temp_img_data_pair.raw_image_data))
                                options_instance.save()

                                # Parse the HTML using BeautifulSoup4 library
                                html_obj = BeautifulSoup(value, 'html.parser')
                                # Find the element with "img" tag
                                my_img_element = html_obj.find('img')
                                my_img_element['src'] = options_instance.image.url  # change src attribute
                                value = str(html_obj)  # save html as string
                                options_instance.text = value  # update options text field
                                options_instance.save()  # save/update entry in database

                elif the_question_type == 'true_false_question':
                    node = node.find('.//response_lid')
                    answer_choices_dict = {}
                    for response_label_elem in node.findall('.//response_label'):
                        response_ident = response_label_elem.get('ident')
                        response_text = response_label_elem.find('.//mattext').text
                        answer_choices_dict[response_ident] = response_text

                    node = item.find('resprocessing')
                    for respcondition_elem in node.findall('.//respcondition'):
                        if respcondition_elem.get('continue') == "No":
                            temp_node = respcondition_elem.find('.//varequal')
                            correct_answer_ident = temp_node.text

                    question_instance = create_question(the_course, the_question_type, question_text_field,
                                                        max_points_for_question)
                    # this block of code checks for an embedded graphic in text, then saves it to field
                    # is there is a graphic. otherwise it does nothing.
                    image_data_pair = check_embedded_graphic(question_text_field)
                    if image_data_pair is not None:
                        # Save the image to the img field, then update the record/entry
                        question_instance.img.save(image_data_pair.actual_image_name,
                                                   ContentFile(image_data_pair.raw_image_data))
                        question_instance.save()
                        print(f"{question_instance.img.url}")

                        # Parse the HTML using BeautifulSoup4 library
                        html_obj = BeautifulSoup(question_text_field, 'html.parser')
                        # Find the element with "img" tag
                        my_img_element = html_obj.find('img')
                        my_img_element['src'] = question_instance.img.url  # change src attribute
                        question_text_field = str(html_obj)  # save html as string
                        question_instance.text = question_text_field  # update text field
                        question_instance.save()  # save/update entry in database

                    testquestion_instance = TestQuestion.objects.create(
                        test=test_instance,
                        question=question_instance,
                        assigned_points=max_points_for_question,
                        section=test_section_instance
                    )

                    for key, value in answer_choices_dict.items():
                        if key == correct_answer_ident:
                            question_instance.answer = value
                            temp_img_data_pair = check_embedded_graphic(value)
                            if temp_img_data_pair is not None:
                                question_instance.ansimg.save(temp_img_data_pair.actual_image_name,
                                                              ContentFile(temp_img_data_pair.raw_image_data))

                            question_instance.save()

                elif the_question_type == 'short_answer_question':  # fill-in-the-blank question (single)

                    the_question_type = "fill_in_the_blank"

                    question_instance = create_question(
                        the_course, the_question_type, question_text_field,
                        max_points_for_question
                    )
                    # this block of code checks for an embedded graphic in text, then saves it to field
                    # is there is a graphic. otherwise it does nothing.
                    image_data_pair = check_embedded_graphic(question_text_field)
                    if image_data_pair is not None:
                        # Save the image to the img field, then update the record/entry
                        question_instance.img.save(image_data_pair.actual_image_name,
                                                   ContentFile(image_data_pair.raw_image_data))
                        question_instance.save()
                        print(f"{question_instance.img.url}")

                        # Parse the HTML using BeautifulSoup4 library
                        html_obj = BeautifulSoup(question_text_field, 'html.parser')
                        # Find the element with "img" tag
                        my_img_element = html_obj.find('img')
                        my_img_element['src'] = question_instance.img.url  # change src attribute
                        question_text_field = str(html_obj)  # save html as string
                        question_instance.text = question_text_field  # update text field
                        question_instance.save()  # save/update entry in database

                    testquestion_instance = TestQuestion.objects.create(
                        test=test_instance,
                        question=question_instance,
                        assigned_points=max_points_for_question,
                        section=test_section_instance
                    )

                    node = item.find('resprocessing')
                    for respcondition_elem in node.findall('.//respcondition'):
                        if respcondition_elem.get('continue') == "No":
                            for varequal_elem in respcondition_elem.findall('.//varequal'):
                                answer_instance = Answers.objects.create(
                                    question=question_instance,
                                    text=varequal_elem.text
                                )
                                temp_img_data_pair = check_embedded_graphic(varequal_elem.text)
                                if temp_img_data_pair is not None:
                                    answer_instance.answer_graphic.save(temp_img_data_pair.actual_image_name,
                                                                        ContentFile(temp_img_data_pair.raw_image_data))
                                    answer_instance.save()

                elif the_question_type == 'multiple_answers_question':

                    the_question_type = "multiple_selection"

                    correct_answer_ident_list = []
                    node = node.find('.//response_lid')
                    answer_choices_dict = {}
                    for response_label_elem in node.findall('.//response_label'):
                        response_ident = response_label_elem.get('ident')
                        response_text = response_label_elem.find('.//mattext').text
                        answer_choices_dict[response_ident] = response_text

                    node = item.find('resprocessing')
                    for respcondition_elem in node.findall('.//respcondition'):
                        if respcondition_elem.get('continue') == "No":
                            for varequal_elem in respcondition_elem.find('conditionvar').find('and').findall(
                                    'varequal'):
                                correct_answer_ident_list.append(varequal_elem.text)

                    question_instance = create_question(the_course, the_question_type, question_text_field,
                                                        max_points_for_question)
                    # this block of code checks for an embedded graphic in text, then saves it to field
                    # is there is a graphic. otherwise it does nothing.
                    image_data_pair = check_embedded_graphic(question_text_field)
                    if image_data_pair is not None:
                        # Save the image to the img field, then update the record/entry
                        question_instance.img.save(image_data_pair.actual_image_name,
                                                   ContentFile(image_data_pair.raw_image_data))
                        question_instance.save()
                        print(f"{question_instance.img.url}")

                        # Parse the HTML using BeautifulSoup4 library
                        html_obj = BeautifulSoup(question_text_field, 'html.parser')
                        # Find the element with "img" tag
                        my_img_element = html_obj.find('img')
                        my_img_element['src'] = question_instance.img.url  # change src attribute
                        question_text_field = str(html_obj)  # save html as string
                        question_instance.text = question_text_field  # update text field
                        question_instance.save()  # save/update entry in database

                    testquestion_instance = TestQuestion.objects.create(
                        test=test_instance,
                        question=question_instance,
                        assigned_points=max_points_for_question,
                        section=test_section_instance
                    )

                    for key, value in answer_choices_dict.items():
                        if key in correct_answer_ident_list:
                            answer_instance = Answers.objects.create(
                                question=question_instance,
                                text=value
                            )
                            temp_img_data_pair = check_embedded_graphic(value)
                            if temp_img_data_pair is not None:
                                answer_instance.answer_graphic.save(temp_img_data_pair.actual_image_name,
                                                                    ContentFile(temp_img_data_pair.raw_image_data))
                                answer_instance.save()

                                # Parse the HTML using BeautifulSoup4 library
                                html_obj = BeautifulSoup(value, 'html.parser')
                                # Find the element with "img" tag
                                my_img_element = html_obj.find('img')
                                my_img_element['src'] = answer_instance.answer_graphic.url  # change src attribute
                                value = str(html_obj)  # save html as string
                                answer_instance.text = value  # update answer_instance text field
                                answer_instance.save()  # save/update entry in database

                        else:
                            options_instance = Options.objects.create(
                                question=question_instance,
                                text=value
                            )
                            temp_img_data_pair = check_embedded_graphic(value)
                            if temp_img_data_pair is not None:
                                options_instance.image.save(temp_img_data_pair.actual_image_name,
                                                            ContentFile(temp_img_data_pair.raw_image_data))
                                options_instance.save()

                                # Parse the HTML using BeautifulSoup4 library
                                html_obj = BeautifulSoup(value, 'html.parser')
                                # Find the element with "img" tag
                                my_img_element = html_obj.find('img')
                                my_img_element['src'] = options_instance.image.url  # change src attribute
                                value = str(html_obj)  # save html as string
                                options_instance.text = value  # update option text field
                                options_instance.save()  # save/update entry in database

                elif the_question_type == 'matching_question':  # this is explicitly stated in rubric to support
                    # Canvas requires you to add at least one answer
                    answer_choices_dict = {}  # right side options and their ID's
                    left_side_dict = {}

                    question_instance = create_question(the_course, the_question_type, question_text_field,
                                                        max_points_for_question)
                    # this block of code checks for an embedded graphic in text, then saves it to field
                    # is there is a graphic. otherwise it does nothing.
                    image_data_pair = check_embedded_graphic(question_text_field)
                    if image_data_pair is not None:
                        # Save the image to the img field, then update the record/entry
                        question_instance.img.save(image_data_pair.actual_image_name,
                                                   ContentFile(image_data_pair.raw_image_data))
                        question_instance.save()
                        print(f"{question_instance.img.url}")

                        # Parse the HTML using BeautifulSoup4 library
                        html_obj = BeautifulSoup(question_text_field, 'html.parser')
                        # Find the element with "img" tag
                        my_img_element = html_obj.find('img')
                        my_img_element['src'] = question_instance.img.url  # change src attribute
                        question_text_field = str(html_obj)  # save html as string
                        question_instance.text = question_text_field  # update text field
                        question_instance.save()  # save/update entry in database

                    testquestion_instance = TestQuestion.objects.create(
                        test=test_instance,
                        question=question_instance,
                        assigned_points=max_points_for_question,
                        section=test_section_instance
                    )

                    # find left sides
                    for response_lid_elem in node.findall('response_lid'):  # for all response_lid elements in list
                        side_key = response_lid_elem.get('ident')
                        side_text = response_lid_elem.find('material').find('mattext').text
                        left_side_dict[side_key] = side_text
                    # find right sides
                    for response_label_elem in node.find('response_lid').find('render_choice').findall(
                            'response_label'):
                        side_key = response_label_elem.get('ident')
                        side_text = response_label_elem.find('.//mattext').text
                        answer_choices_dict[side_key] = side_text
                    # now find out which ones are matching pairs
                    matching_pairs_dict = {}
                    node = item.find('resprocessing')
                    right_side_key_to_delete_list = []
                    for respcondition_elem in node.findall('respcondition'):
                        varequal_elem = respcondition_elem.find('conditionvar').find('varequal')
                        if varequal_elem is not None:
                            left_key = varequal_elem.get('respident')
                            right_key = varequal_elem.text
                            right_side_key_to_delete_list.append(right_key)
                            # this makes a dictionary of matching pairs
                            matching_pairs_dict[left_side_dict.get(left_key)] = answer_choices_dict.get(right_key)
                    unique_key_list_to_del = list(
                        set(right_side_key_to_delete_list))  # this removes duplicate keys from list
                    for key_string in unique_key_list_to_del:
                        del answer_choices_dict[key_string]  # deletes a response option that was a correct right side
                    # now save matching pairs to database
                    for key, value in matching_pairs_dict.items():
                        matching_pair_string = ""
                        if key is not None:
                            matching_pair_string += key
                        matching_pair_string += ";;;;; "
                        if value is None:
                            matching_pair_string += ";;;;;"
                        else:
                            matching_pair_string += value

                        # matching questions CANNOT have embedded graphics in responses
                        answer_instance = Answers.objects.create(
                            question=question_instance,
                            text=matching_pair_string
                        )
                    # save distractors to database
                    for value in answer_choices_dict.values():
                        option_instance = Options.objects.create(
                            question=question_instance,
                            text=value
                        )

                elif the_question_type == 'essay_question':
                    # mostly done but may need to process feedbacks or comments

                    # this creates a question record in database
                    question_instance = create_question(the_course, the_question_type, question_text_field,
                                                        max_points_for_question)
                    # this block of code checks for an embedded graphic in text, then saves it to field
                    # is there is a graphic. otherwise it does nothing.
                    image_data_pair = check_embedded_graphic(question_text_field)
                    if image_data_pair is not None:
                        # Save the image to the img field, then update the record/entry
                        question_instance.img.save(image_data_pair.actual_image_name,
                                                   ContentFile(image_data_pair.raw_image_data))
                        question_instance.save()
                        print(f"{question_instance.img.url}")

                        # Parse the HTML using BeautifulSoup4 library
                        html_obj = BeautifulSoup(question_text_field, 'html.parser')
                        # Find the element with "img" tag
                        my_img_element = html_obj.find('img')
                        my_img_element['src'] = question_instance.img.url  # change src attribute
                        question_text_field = str(html_obj)  # save html as string
                        question_instance.text = question_text_field  # update text field
                        question_instance.save()  # save/update entry in database

                    testquestion_instance = TestQuestion.objects.create(
                        test=test_instance,
                        question=question_instance,
                        assigned_points=max_points_for_question,
                        section=test_section_instance
                    )

                # commented out because currently not supported
                """
                elif the_question_type == 'fill_in_multiple_blanks_question':
                    print('')  # this is placeholder for code to extract info from question for table
                elif the_question_type == 'multiple_dropdowns_question':
                    print('')  # this is placeholder for code to extract info from question for table
                elif the_question_type == 'numerical_question':
                    print('')  # this is placeholder for code to extract info from question for table
                elif the_question_type == 'calculated_question':
                    print('')  # this is placeholder for code to extract info from question for table

                elif the_question_type == 'file_upload_question':
                    # should be done but may need to process feedbacks or comments
                    print('')
                elif the_question_type == 'text_only_question':
                    # placeholder for any future changes, but 99.9% sure this is done
                    print('')
                """

                #

    # file_info is just used for testing. remove after (probably)
    # for now, what the Parser returns depends on this
    file_info = None

    uploaded_file = None

    # """
    # 00 Begin
    # Code in triple quotes is used for when merged with frontend
    # "file" in request.FILES.get("file") changes or depends on something in the HTML/javascript form
    if request.method == "POST" and request.FILES.get("file"):
        uploaded_file = request.FILES["file"]  # Get the uploaded file
        print("File uploaded:", uploaded_file.name)

        file_info = {
            "filename": uploaded_file.name,
            "size": uploaded_file.size
        }
    else:
        print("No file uploaded to website.")

    if uploaded_file is None:
        return JsonResponse({"message": "No file uploaded or it doesn't exist.", "file_info": file_info})

    course_id = request.POST.get("courseID")
    course_name = request.POST.get("courseName")
    course_crn = request.POST.get("courseCRN")
    course_semester = request.POST.get("courseSemester")
    course_textbook_title = request.POST.get("courseTextbookTitle")
    course_textbook_author = request.POST.get("courseTextbookAuthor")
    course_textbook_version = request.POST.get("courseTextbookVersion")
    course_textbook_isbn = request.POST.get("courseTextbookISBN")
    course_textbook_link = request.POST.get("courseTextbookLink")

    textbook_instance, created = Textbook.objects.get_or_create(
        title=course_textbook_title,
        author=course_textbook_author,
        version=course_textbook_version,
        isbn=course_textbook_isbn,
        defaults={
            "link": course_textbook_link,
        }
    )

    course_instance, created = Course.objects.get_or_create(
        course_id=course_id,
        defaults={
            "name": course_name,
            "crn": course_crn,
            "sem": course_semester,
            "textbook": textbook_instance
        }
    )

    # Check if the user is authenticated (logged in)
    if request.user.is_authenticated:

        course_instance.user = request.user  # sets field to current user
        course_instance.save()

        current_user = request.user
        # Check if teacher already in course
        if current_user in course_instance.teachers.all():
            print(f'{current_user.username} teacher already in {course_instance.name} course')
        else:
            course_instance.teachers.add(current_user)  # Adds teacher if not in course
            course_instance.save()  # Updates the Course entry in the database (makes sure it's saved)
            print(f'{current_user.username} teacher ADDED to {course_instance.name} course')
    else:
        print("User is not logged in.")

    # 00 End
    # """

    # this is used to stop removing and adding "#" when switching between tests
    if uploaded_file is None:
        path_to_zip_file = 'qti sample w one quiz-slash-test w all typesofquestions.zip'
        # path_to_zip_file = 'added response feedback.zip'

        # Get the first available course (REMOVE AFTER TESTING)
        course_instance = Course.objects.first()
        # Used for testing. Remove after
        if course_instance is None:
            course_instance = Course.objects.create(
                course_code='CS123',
                course_name='placeholder_course',
                textbook_title='placeholder Tb title',
                textbook_author='placeholder author',
                textbook_isbn='placeholder isbn',
                textbook_link='placeholder Tb link'
            )
    else:
        path_to_zip_file = uploaded_file

    with zipfile.ZipFile(path_to_zip_file, 'r') as zip_ref:
        # List all files inside the zip file
        filename_list = zip_ref.namelist()

        for file_name in filename_list:

            # looks for folders that are direct children of the zipfile
            if file_name.endswith('/'):

                temp_file_list = []

                # looks for files that are children of the found folder, then adds them to a list
                for temp_filename in filename_list:
                    if temp_filename.startswith(f'{file_name}') and (temp_filename != file_name):
                        temp_file_list.append(temp_filename)

                # if a folder is not empty, process the files in it with the parser
                if temp_file_list and len(temp_file_list) >= 2 and temp_file_list[0].endswith('.xml') and \
                        temp_file_list[1].endswith('.xml'):
                    # sort the files in the list because the metadata file for assessments is always
                    # named assessment_meta.xml, but the file with questions seems to always
                    # start with the letter "g"
                    temp_file_list = sorted(temp_file_list)
                    assessment_meta_path = temp_file_list[0]
                    questions_file_path = temp_file_list[1]

                    with zip_ref.open(assessment_meta_path) as outer_file:
                        with zip_ref.open(questions_file_path) as inner_file:
                            outer_file.seek(0)  # Reset file pointer
                            inner_file.seek(0)

                            # this calls the function that actually handles the parsing
                            parse_just_xml(outer_file, inner_file, course_instance)
                            #

    #
    end_time = time.perf_counter()
    execution_time = end_time - start_time
    print(execution_time)
    # this is here because the javascript that calls the Parser depends on what it returns
    if file_info is None:
        print("Success! Created test record")
        return JsonResponse({"Success": "created Test record."}, status=555)

    else:
        print("File processed successfully!")
        return JsonResponse({"message": "File processed successfully!", "file_info": file_info})

#
